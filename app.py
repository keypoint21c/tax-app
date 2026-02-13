# app.py
# ==========================================================
# 개인사업자 성실신고 리스크 & 법인전환 전략 분석 (Streamlit)
# - Supabase 승인/로그인(이메일만) + 관리자 승인관리 화면 포함
# - 엑셀 업로드 기반 소득율 자동 계산(F->C, K->Q, 소득율=100-Q)
#
# 필요한 Streamlit Secrets (TOML):
# SUPABASE_URL = "https://xxxxx.supabase.co"
# SUPABASE_SERVICE_ROLE_KEY = "서비스 롤 키(절대 공개X)"
# ADMIN_BOOTSTRAP_KEY = "아주긴랜덤문자열"
#
# (선택) 기본 관리자 이메일을 하드코딩하고 싶으면 DEFAULT_ADMIN_EMAIL 사용
# DEFAULT_ADMIN_EMAIL = "keypoint21c@gmail.com"
#
# requirements.txt 예시:
# streamlit
# pandas
# openpyxl
# supabase
# ==========================================================

import os
import math
from dataclasses import dataclass
from typing import Optional, Dict, Any, Tuple, List

import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# Supabase (supabase-py)
try:
    from supabase import create_client
except Exception:
    create_client = None


# -----------------------------
# Streamlit 기본 설정
# -----------------------------
st.set_page_config(page_title="성실신고 리스크 & 법인전환 전략 분석", layout="wide")


# -----------------------------
# 유틸: 숫자 표시
# -----------------------------
def fmt_won(x: float) -> str:
    try:
        return f"{int(round(x)):,.0f}원"
    except Exception:
        return f"{x}원"


def fmt_pct(x: float, nd: int = 2) -> str:
    try:
        return f"{x:.{nd}f}%"
    except Exception:
        return f"{x}%"


# -----------------------------
# (핵심) 엑셀에서 소득율 산출
#   - F열: 산업분류코드
#   - C열: 업종코드
#   - K열: 업종코드
#   - Q열: Q값
#   - 소득율 = 100 - Q값
# -----------------------------
@dataclass
class IncomeRateResult:
    industry_code: str
    biz_code: str
    q_value: float
    income_rate_pct: float


def excel_col_letter_to_index(letter: str) -> int:
    # A=1, B=2 ... (openpyxl 기준)
    letter = letter.strip().upper()
    n = 0
    for ch in letter:
        if not ("A" <= ch <= "Z"):
            raise ValueError("Invalid column letter")
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


@st.cache_data(show_spinner=False)
def load_workbook_cached(file_bytes: bytes):
    # openpyxl은 파일 객체를 필요로 하므로 bytes를 temp로 처리
    # streamlit cache에는 bytes->wb 반환 형태로 저장
    from io import BytesIO
    bio = BytesIO(file_bytes)
    wb = load_workbook(bio, data_only=True)
    return wb


def find_value_in_column(ws, col_letter: str, target: str) -> Optional[int]:
    """지정 열(col_letter)에서 target과 '문자열 기준으로 동일'한 행 번호를 찾는다."""
    col_idx = excel_col_letter_to_index(col_letter)
    target_norm = str(target).strip()

    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is None:
            continue
        if str(v).strip() == target_norm:
            return r
    return None


def read_cell(ws, col_letter: str, row: int):
    col_idx = excel_col_letter_to_index(col_letter)
    return ws.cell(row=row, column=col_idx).value


def compute_income_rate_from_excel(file_bytes: bytes, industry_code: str) -> IncomeRateResult:
    wb = load_workbook_cached(file_bytes)
    ws = wb.active

    # 1) F열에서 산업분류코드 찾기 -> 그 행의 C열 = 업종코드
    row_f = find_value_in_column(ws, "F", industry_code)
    if row_f is None:
        raise ValueError(f"엑셀 F열에서 산업분류코드({industry_code})를 찾지 못했습니다.")

    biz_code = read_cell(ws, "C", row_f)
    if biz_code is None or str(biz_code).strip() == "":
        raise ValueError("해당 행의 C열(업종코드)이 비어 있습니다.")
    biz_code_str = str(biz_code).strip()

    # 2) K열에서 업종코드 찾기 -> 그 행의 Q열 = Q값
    row_k = find_value_in_column(ws, "K", biz_code_str)
    if row_k is None:
        raise ValueError(f"엑셀 K열에서 업종코드({biz_code_str})를 찾지 못했습니다.")

    q_val = read_cell(ws, "Q", row_k)
    if q_val is None or str(q_val).strip() == "":
        raise ValueError("해당 행의 Q열(Q값)이 비어 있습니다.")

    try:
        q_val_f = float(q_val)
    except Exception:
        raise ValueError(f"Q값이 숫자가 아닙니다: {q_val}")

    income_rate = 100.0 - q_val_f
    return IncomeRateResult(
        industry_code=str(industry_code).strip(),
        biz_code=biz_code_str,
        q_value=q_val_f,
        income_rate_pct=income_rate,
    )


# -----------------------------
# 세금(종합소득세) 계산 (단순화 버전)
# - 실제 공제/필요경비/세액공제는 반영 안됨
# - "리스크 체감" 목적의 추정치
# -----------------------------
# (참고) 2024년 기준으로 널리 알려진 누진 구간(단순 적용).
# 만약 최신 세율/구간이 변경되면 아래만 수정하면 됨.
INCOME_TAX_BRACKETS = [
    (14_000_000, 0.06),
    (50_000_000, 0.15),
    (88_000_000, 0.24),
    (150_000_000, 0.35),
    (300_000_000, 0.38),
    (500_000_000, 0.40),
    (1_000_000_000, 0.42),
    (float("inf"), 0.45),
]


def calc_progressive_tax(taxable: float) -> float:
    """누진세(단순) 계산: 과세표준을 taxable로 보고 구간별 누진 계산"""
    if taxable <= 0:
        return 0.0

    tax = 0.0
    prev = 0.0
    for limit, rate in INCOME_TAX_BRACKETS:
        if taxable <= limit:
            tax += (taxable - prev) * rate
            break
        tax += (limit - prev) * rate
        prev = limit
    return tax


def calc_total_income_tax_with_local(taxable: float, local_rate: float = 0.10) -> float:
    nat = calc_progressive_tax(taxable)
    local = nat * local_rate
    return nat + local


# -----------------------------
# 성실신고확인대상 위험도 판단
# -----------------------------
def sungshil_risk_level(category: str, sales: float) -> Tuple[str, float]:
    """
    category:
      - 도소매: 15억 이상
      - 제조/건설: 7.5억 이상
      - 서비스/임대: 5억 이상
    return: (위험도 라벨, 기준값)
    """
    cat = category.strip()
    if cat == "도소매":
        threshold = 1_500_000_000
    elif cat == "제조/건설":
        threshold = 750_000_000
    else:  # 서비스/임대
        threshold = 500_000_000

    ratio = sales / threshold if threshold > 0 else 0

    if ratio < 0.7:
        return ("낮음", threshold)
    if ratio < 1.0:
        return ("보통", threshold)
    if ratio < 1.3:
        return ("높음", threshold)
    return ("매우 높음", threshold)


# -----------------------------
# Supabase DB (users 테이블)
# users(email text primary key, approved boolean, is_admin boolean)
# -----------------------------
def get_secret(name: str, default: Optional[str] = None) -> Optional[str]:
    # Streamlit Cloud: st.secrets
    # 로컬: 환경변수 fallback
    try:
        if name in st.secrets:
            return str(st.secrets[name])
    except Exception:
        pass
    return os.getenv(name, default)


def get_supabase_client():
    if create_client is None:
        st.error("Supabase 라이브러리(supabase)가 설치되지 않았습니다. requirements.txt에 'supabase'를 추가하세요.")
        st.stop()

    url = get_secret("SUPABASE_URL")
    key = get_secret("SUPABASE_SERVICE_ROLE_KEY")  # 서버에서만 쓰는 키 (절대 유출 X)

    if not url or not key:
        st.error("Supabase 설정이 없습니다. Streamlit Cloud → Settings → Secrets에 SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY를 넣어주세요.")
        st.stop()

    return create_client(url, key)


def db_get_user(sb, email: str) -> Optional[Dict[str, Any]]:
    res = sb.table("users").select("*").eq("email", email).execute()
    data = getattr(res, "data", None)
    if not data:
        return None
    return data[0]


def db_upsert_user(sb, email: str, approved: bool = False, is_admin: bool = False) -> Dict[str, Any]:
    payload = {"email": email, "approved": approved, "is_admin": is_admin}
    res = sb.table("users").upsert(payload).execute()
    data = getattr(res, "data", None)
    if not data:
        # upsert 후 select로 재조회
        u = db_get_user(sb, email)
        if u:
            return u
        raise RuntimeError("DB upsert 실패")
    return data[0]


def db_set_approval(sb, email: str, approved: bool):
    sb.table("users").update({"approved": approved}).eq("email", email).execute()


def db_set_admin(sb, email: str, is_admin: bool):
    sb.table("users").update({"is_admin": is_admin}).eq("email", email).execute()


def db_list_users(sb) -> List[Dict[str, Any]]:
    res = sb.table("users").select("*").order("email").execute()
    return getattr(res, "data", []) or []


# -----------------------------
# 로그인/세션
# -----------------------------
def session_get_email() -> Optional[str]:
    return st.session_state.get("auth_email")


def session_set_email(email: Optional[str]):
    st.session_state["auth_email"] = email


def normalize_email(email: str) -> str:
    return email.strip().lower()


def ensure_user_record(sb, email: str) -> Dict[str, Any]:
    email = normalize_email(email)
    u = db_get_user(sb, email)
    if u is None:
        # 최초 접속자는 승인 false로 등록
        u = db_upsert_user(sb, email, approved=False, is_admin=False)

        # (선택) DEFAULT_ADMIN_EMAIL로 자동 관리자 등록
        default_admin = get_secret("DEFAULT_ADMIN_EMAIL")
        if default_admin and normalize_email(default_admin) == email:
            u = db_upsert_user(sb, email, approved=True, is_admin=True)

    return u


# -----------------------------
# UI: 사이드바 로그인/접근제어
# -----------------------------
def render_access_sidebar(sb):
    st.sidebar.markdown("## 🔐 접근 제어")

    cur_email = session_get_email()
    if cur_email:
        st.sidebar.success(f"로그인됨: {cur_email}")
        if st.sidebar.button("로그아웃"):
            session_set_email(None)
            st.rerun()
        st.sidebar.divider()
        return

    email = st.sidebar.text_input("이메일", placeholder="name@example.com")
    if st.sidebar.button("로그인", use_container_width=True):
        if not email or "@" not in email:
            st.sidebar.error("이메일을 올바르게 입력하세요.")
        else:
            email_n = normalize_email(email)
            _ = ensure_user_record(sb, email_n)
            session_set_email(email_n)
            st.rerun()

    st.sidebar.caption("※ 최초 로그인 시 DB에 자동 등록되며, 관리자가 승인하면 사용 가능합니다.")


# -----------------------------
# UI: 관리자 패널
# -----------------------------
def render_admin_panel(sb, me: Dict[str, Any]):
    st.subheader("🛠 관리자 기능")

    # 1) 관리자 초기설정(부트스트랩)
    with st.expander("🛠 관리자 초기설정(최초 1회)", expanded=False):
        st.write("ADMIN_BOOTSTRAP_KEY가 맞으면, 현재 로그인 이메일을 관리자/승인 처리합니다.")
        boot = st.text_input("ADMIN_BOOTSTRAP_KEY", type="password")
        if st.button("관리자 계정 생성/갱신"):
            expected = get_secret("ADMIN_BOOTSTRAP_KEY")
            if not expected:
                st.error("Secrets에 ADMIN_BOOTSTRAP_KEY가 없습니다.")
            elif boot != expected:
                st.error("키가 일치하지 않습니다.")
            else:
                db_upsert_user(sb, me["email"], approved=True, is_admin=True)
                st.success("관리자/승인 처리 완료! 앱을 새로고침하세요.")
                st.rerun()

    st.divider()

    # 2) 승인 관리 테이블
    st.markdown("### ✅ 승인 관리")
    users = db_list_users(sb)
    if not users:
        st.info("users 테이블에 데이터가 없습니다.")
        return

    df = pd.DataFrame(users)
    # 보기 편하게 정렬/표시
    df = df[["email", "approved", "is_admin"]].sort_values("email")

    st.dataframe(df, use_container_width=True, hide_index=True)

    st.markdown("#### 승인/관리자 변경")
    col1, col2, col3 = st.columns([2, 1, 1])
    target_email = col1.text_input("대상 이메일", placeholder="someone@example.com")
    new_approved = col2.selectbox("승인", options=[True, False], index=0)
    new_admin = col3.selectbox("관리자", options=[True, False], index=1)

    if st.button("변경 적용"):
        if not target_email or "@" not in target_email:
            st.error("대상 이메일이 올바르지 않습니다.")
        else:
            e = normalize_email(target_email)
            ensure_user_record(sb, e)
            db_set_approval(sb, e, bool(new_approved))
            db_set_admin(sb, e, bool(new_admin))
            st.success("변경 완료")
            st.rerun()


# -----------------------------
# 보고서 계산
# -----------------------------
@dataclass
class ReportInput:
    industry_code: str
    last_sales: float
    this_sales: float
    employees: int
    category: str

    health_rate: float
    corp_tax_rate: float
    ceo_salary: float

    # 비용 부인 가정 비율
    deny_outsource: float
    deny_family_pay: float
    deny_private: float
    deny_cash: float


def build_report(inp: ReportInput, income_rate_pct: float) -> Dict[str, Any]:
    rate = income_rate_pct / 100.0

    # 1) 순이익 추정
    last_profit = inp.last_sales * rate
    this_profit = inp.this_sales * rate

    # 2) 종합소득세(지방세 포함) 추정
    tax_this = calc_total_income_tax_with_local(this_profit)
    tax_last = calc_total_income_tax_with_local(last_profit)

    # 민감도 (+/- 1%p 소득율)
    tax_this_up = calc_total_income_tax_with_local(inp.this_sales * ((income_rate_pct + 1.0) / 100.0))
    tax_this_dn = calc_total_income_tax_with_local(inp.this_sales * ((income_rate_pct - 1.0) / 100.0))
    delta_up = tax_this_up - tax_this
    delta_dn = tax_this - tax_this_dn

    # 3) 성실신고 리스크
    risk_label, threshold = sungshil_risk_level(inp.category, inp.this_sales)

    # 4) 비용 부인 시뮬레이션
    deny_items = [
        ("외주가공비(부인)", inp.deny_outsource),
        ("가족·특수관계인 인건비(부인)", inp.deny_family_pay),
        ("차량·접대 등 사적경비(부인)", inp.deny_private),
        ("무증빙·현금지출(부인)", inp.deny_cash),
    ]

    rows = []
    base_tax = tax_this
    total_deny = 0.0
    total_add_tax = 0.0
    total_add_health = 0.0

    for name, r in deny_items:
        deny_amt = inp.this_sales * r
        new_tax = calc_total_income_tax_with_local(this_profit + deny_amt)
        add_tax = new_tax - base_tax
        add_health = deny_amt * inp.health_rate

        rows.append({
            "항목": name,
            "가정 부인금액": deny_amt,
            "증가 과세소득": deny_amt,
            "추가 종합소득세(지방세 포함)": add_tax,
            "건강보험 증가 추정": add_health,
        })

        total_deny += deny_amt
        total_add_tax += add_tax
        total_add_health += add_health

    sim_df = pd.DataFrame(rows)

    # 5) 3년 누적
    base_3y = (base_tax + (this_profit * inp.health_rate)) * 3
    after_3y = ((base_tax + total_add_tax) + ((this_profit + total_deny) * inp.health_rate)) * 3
    inc_3y = after_3y - base_3y

    # 6) 법인 전환 비교(단순화)
    # - 대표 급여는 비용으로 처리된다고 가정(법인 과세표준 감소)
    corp_taxable = max(0.0, this_profit - inp.ceo_salary)
    corp_tax = corp_taxable * inp.corp_tax_rate

    # 개인 유지/성실신고 정리/법인 전환 3년 비교(아주 단순)
    # 개인 유지: base_tax + 건강(이익*rate) *3
    # 성실신고 정리: (base_tax+add_tax) + 건강((이익+부인)*rate) *3
    # 법인 전환: 법인세 + (대표 급여에 대한 개인세는 미반영) + 건강(직장 전환 효과는 '절감'으로 표현만)
    corp_3y = corp_tax * 3  # 단순 (추가로 4대보험/급여 소득세 등은 별도)
    compare = pd.DataFrame([
        {"구분": "개인 유지", "3년 추정세금+건보(단순)": base_3y},
        {"구분": "성실신고 정리 후", "3년 추정세금+건보(단순)": after_3y},
        {"구분": "법인 전환(법인세 중심 단순)", "3년 추정세금+건보(단순)": corp_3y},
    ])

    # 결론용 문구
    if total_deny > 0:
        per_100m = (total_add_tax / total_deny) if total_deny else 0
        example_text = f"비용 {fmt_won(100_000_000)} 정리 시 세금은 대략 {fmt_won(100_000_000 * per_100m)} 수준으로 증가할 수 있습니다(단순 추정)."
    else:
        example_text = "비용 부인 시뮬레이션 값이 0이라 예시 문구를 만들 수 없습니다."

    return {
        "last_profit": last_profit,
        "this_profit": this_profit,
        "tax_last": tax_last,
        "tax_this": base_tax,
        "delta_up": delta_up,
        "delta_dn": delta_dn,
        "risk_label": risk_label,
        "risk_threshold": threshold,
        "sim_df": sim_df,
        "total_deny": total_deny,
        "total_add_tax": total_add_tax,
        "total_add_health": total_add_health,
        "base_3y": base_3y,
        "after_3y": after_3y,
        "inc_3y": inc_3y,
        "corp_taxable": corp_taxable,
        "corp_tax": corp_tax,
        "compare_df": compare,
        "example_text": example_text,
    }


# -----------------------------
# 메인 UI
# -----------------------------
def main():
    st.title("📊 개인사업자 성실신고 리스크 & 법인전환 전략 분석 (배포용)")

    # Supabase 연결
    sb = get_supabase_client()

    # 사이드바 로그인/접근제어
    render_access_sidebar(sb)

    # 로그인 체크
    email = session_get_email()
    if not email:
        st.info("왼쪽 사이드바에서 이메일로 로그인하세요.")
        st.stop()

    # 유저 상태 조회
    me = ensure_user_record(sb, email)

    # 승인 여부 체크
    if not me.get("approved", False):
        st.warning("등록되었습니다. 관리자 승인 대기 중입니다.")
        st.caption("관리자에게 승인 요청 후 다시 접속하세요.")
        st.stop()

    # 관리자면 패널 표시
    if me.get("is_admin", False):
        with st.expander("🛠 관리자 패널(승인/관리자 설정)", expanded=False):
            render_admin_panel(sb, me)

    st.divider()

    # --------------------------------
    # 입력 UI
    # --------------------------------
    st.subheader("1) 기본 입력")
    c1, c2, c3, c4 = st.columns([1, 1, 1, 1])

    industry_code = c1.text_input("산업분류코드(F열)", value="25913")
    last_sales = c2.number_input("작년 매출(원)", min_value=0, value=800_000_000, step=10_000_000)
    this_sales = c3.number_input("금년 예상 매출(원)", min_value=0, value=1_000_000_000, step=10_000_000)
    employees = int(c4.number_input("직원 수(대표 제외)", min_value=0, value=6, step=1))

    st.subheader("2) 성실신고 기준(업종 분류)")
    category = st.selectbox("업종 분류 선택", options=["제조/건설", "도소매", "서비스/임대"], index=0)

    st.subheader("3) 엑셀 업로드(필수)")
    st.caption("업종코드-표준산업분류 연계표 엑셀을 업로드해야 소득율을 계산할 수 있습니다.")
    xlsx = st.file_uploader("연계표_기준경비율 엑셀(.xlsx) 업로드", type=["xlsx"])

    st.subheader("4) 건보/법인 가정값")
    cc1, cc2, cc3 = st.columns([1, 1, 1])
    health_rate = cc1.slider("건강보험 증가 추정률(과세소득 대비)", 0.00, 0.20, 0.05, 0.01)
    corp_tax_rate = cc2.slider("법인세(단순 가정)", 0.05, 0.25, 0.09, 0.01)
    ceo_salary = cc3.number_input("법인 전환 시 대표 급여 가정(원/년)", min_value=0, value=70_000_000, step=1_000_000)

    st.subheader("5) 성실신고 비용 부인 시뮬레이션(보수적 기본값)")
    s1, s2, s3, s4 = st.columns(4)
    deny_outsource = s1.slider("외주가공비(매출 대비)", 0.0, 0.10, 0.02, 0.005)
    deny_family_pay = s2.slider("가족/특수관계인 인건비(매출 대비)", 0.0, 0.10, 0.01, 0.005)
    deny_private = s3.slider("차량/접대 등 사적경비(매출 대비)", 0.0, 0.10, 0.01, 0.005)
    deny_cash = s4.slider("무증빙/현금지출(매출 대비)", 0.0, 0.05, 0.005, 0.0025)

    st.divider()

    # --------------------------------
    # 보고서 생성 버튼
    # --------------------------------
    if st.button("✅ 보고서 생성", use_container_width=True):
        if not xlsx:
            st.error("엑셀 파일을 업로드해 주세요.")
            st.stop()

        try:
            r = compute_income_rate_from_excel(xlsx.getvalue(), industry_code)
        except Exception as e:
            st.error(f"소득율 산출 실패: {e}")
            st.stop()

        inp = ReportInput(
            industry_code=industry_code,
            last_sales=float(last_sales),
            this_sales=float(this_sales),
            employees=employees,
            category=category,
            health_rate=float(health_rate),
            corp_tax_rate=float(corp_tax_rate),
            ceo_salary=float(ceo_salary),
            deny_outsource=float(deny_outsource),
            deny_family_pay=float(deny_family_pay),
            deny_private=float(deny_private),
            deny_cash=float(deny_cash),
        )

        rep = build_report(inp, r.income_rate_pct)

        # --------------------------------
        # 보고서 출력 (요청한 순서대로)
        # --------------------------------
        st.header("📌 최종 보고서")

        st.subheader("1) 소득율 산출 결과")
        st.write(f"- 산업분류코드: **{r.industry_code}**")
        st.write(f"- 업종코드(C열): **{r.biz_code}**")
        st.write(f"- Q값(Q열): **{r.q_value}**")
        st.write(f"- 계산된 소득율: **{fmt_pct(r.income_rate_pct)}**")

        st.subheader("2) 순이익 추정")
        st.write(f"- 작년 순이익(추정): **{fmt_won(rep['last_profit'])}**")
        st.write(f"- 금년 순이익(추정): **{fmt_won(rep['this_profit'])}**")
        st.caption("※ 순이익=매출×소득율(단순). 실제는 경비/소득구성에 따라 달라집니다.")

        st.subheader("3) 종합소득세 계산(지방소득세 포함, 단순 추정)")
        st.write(f"- 작년 예상 세금: **{fmt_won(rep['tax_last'])}**")
        st.write(f"- 금년 예상 세금: **{fmt_won(rep['tax_this'])}**")
        st.write(f"- 소득율 +1%p 시 세금 증가(추정): **{fmt_won(rep['delta_up'])}**")
        st.write(f"- 소득율 -1%p 시 세금 감소(추정): **{fmt_won(rep['delta_dn'])}**")
        st.caption("※ 공제/세액공제/기타소득 합산 등은 미반영된 ‘리스크 체감용’ 추정치입니다.")

        st.subheader("4) 성실신고확인대상 여부 판단(국세청 기준 기반)")
        st.write(f"- 업종 분류: **{category}**")
        st.write(f"- 기준 매출: **{fmt_won(rep['risk_threshold'])}**")
        st.write(f"- 금년 매출: **{fmt_won(this_sales)}**")
        st.write(f"- 위험도: **{rep['risk_label']}**")

        st.subheader("5) 성실신고 시 비용 부인 시뮬레이션")
        st.dataframe(
            rep["sim_df"].assign(
                **{
                    "가정 부인금액": rep["sim_df"]["가정 부인금액"].map(fmt_won),
                    "증가 과세소득": rep["sim_df"]["증가 과세소득"].map(fmt_won),
                    "추가 종합소득세(지방세 포함)": rep["sim_df"]["추가 종합소득세(지방세 포함)"].map(fmt_won),
                    "건강보험 증가 추정": rep["sim_df"]["건강보험 증가 추정"].map(fmt_won),
                }
            ),
            use_container_width=True,
            hide_index=True
        )

        st.write(f"- 총 비용 부인 금액: **{fmt_won(rep['total_deny'])}**")
        st.write(f"- 총 추가 세금(추정): **{fmt_won(rep['total_add_tax'])}**")
        st.write(f"- 총 건강보험 증가(추정): **{fmt_won(rep['total_add_health'])}**")
        st.info(rep["example_text"])

        st.subheader("6) 3년 누적 리스크 계산(단순)")
        st.write(f"- 현재 구조 유지(3년) 세금+건보(단순): **{fmt_won(rep['base_3y'])}**")
        st.write(f"- 성실신고 비용 정리 후(3년) 세금+건보(단순): **{fmt_won(rep['after_3y'])}**")
        st.write(f"- 3년 증가분(단순): **{fmt_won(rep['inc_3y'])}**")
        st.caption("※ 5년 누적은 변동성이 커서 ‘확률/추세’로만 언급하는 것을 권장합니다.")

        st.subheader("7) 법인 전환 시 비교 분석(단순)")
        st.write(f"- 법인 과세표준(단순): max(0, 순이익 - 대표급여) = **{fmt_won(rep['corp_taxable'])}**")
        st.write(f"- 법인세(단순): 과세표준 × {fmt_pct(inp.corp_tax_rate*100, 2)} = **{fmt_won(rep['corp_tax'])}**")
        st.caption("※ 실제는 대표 급여 소득세/4대보험/업무용승용차/퇴직금/배당 등 설계가 핵심입니다.")

        st.markdown("#### 3년 누적 비교표(단순)")
        cdf = rep["compare_df"].copy()
        cdf["3년 추정세금+건보(단순)"] = cdf["3년 추정세금+건보(단순)"].map(fmt_won)
        st.dataframe(cdf, use_container_width=True, hide_index=True)

        st.subheader("8) 전략적 결론(상담용)")
        st.write(
            "- **핵심 리스크**: 매출 규모가 성실신고 기준에 근접/초과하면 ‘비용 정리(부인)’가 발생할 때 세금과 건보가 동시에 뛰는 구조입니다.\n"
            "- **대응 방향**: (1) 증빙 체계 강화 + (2) 비용 항목 구조 점검 + (3) 법인 전환/급여·배당 설계로 리스크를 분산하는 시나리오를 병행합니다.\n"
            "- **다음 액션**: 실제 계정별 비용/인건비/외주 구조를 받아 ‘부인 가능성’ 높은 항목부터 방어 자료(계약서/작업지시/세금계산서/입금증)를 정리합니다."
        )

        st.subheader("📞 1차 미팅 클로징 멘트(바로 사용)")
        st.write(
            "“대표님, 지금 숫자만 봐도 성실신고 구간에서 **비용 정리 1~2건**이 생기면 "
            "세금과 건보가 **동시에 올라가는 구조**예요. 오늘은 ‘위험이 큰 비용 항목’부터 먼저 잡고, "
            "동시에 **법인 전환/급여 설계 시나리오**까지 같이 비교해서 ‘가장 안전한 선택지’를 만들겠습니다.”"
        )

    st.caption("© 배포용 버전 — 숫자는 ‘리스크 체감용 단순 추정’이며, 실제 신고/설계는 세무사 검토가 필요합니다.")


if __name__ == "__main__":
    main()




